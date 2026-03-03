# -*- coding: utf-8 -*-
import os
import json
import openpyxl
import shutil

excel_path = r"c:\Users\黄文达\Desktop\app\中草药百科与识别后台\药材数据_2026-2-23.xlsx"
herbs_folder = r"c:\Users\黄文达\Desktop\app\中草药百科与识别后台\herbs"
output_path = r"c:\Users\黄文达\Desktop\app\中草药百科与识别后台\药材数据导入.json"

wb = openpyxl.load_workbook(excel_path)
ws = wb.active

headers = []
for col in range(1, ws.max_column + 1):
    headers.append(ws.cell(1, col).value)

print(f"表头: {headers}")

herb_data_list = []
image_name_set = set()

for row in range(2, ws.max_row + 1):
    row_data = {}
    for col in range(1, ws.max_column + 1):
        value = ws.cell(row, col).value
        row_data[headers[col - 1]] = value
    
    herb_name = row_data.get("药材名称", "")
    if herb_name:
        herb_data_list.append(row_data)
        image_name_set.add(herb_name)

print(f"\nExcel 中的药材数量: {len(herb_data_list)}")

existing_images = []
for filename in os.listdir(herbs_folder):
    if filename.endswith(('.jpg', '.jpeg', '.png', '.gif')):
        name_without_ext = os.path.splitext(filename)[0]
        existing_images.append(name_without_ext)

print(f"herbs 文件夹中的图片数量: {len(existing_images)}")

matched = []
not_in_excel = []
not_in_images = []

for img_name in existing_images:
    if img_name in image_name_set:
        matched.append(img_name)
    else:
        not_in_excel.append(img_name)

for herb in herb_data_list:
    herb_name = herb.get("药材名称", "")
    if herb_name not in existing_images:
        not_in_images.append(herb_name)

print(f"\n匹配成功的数量: {len(matched)}")
print(f"图片在 Excel 中找不到: {len(not_in_excel)}")
print(f"Excel 中找不到对应图片: {len(not_in_images)}")

if not_in_images:
    print(f"\nExcel 中找不到对应图片的药材: {not_in_images[:10]}...")

if not_in_excel:
    print(f"\n图片在 Excel 中找不到的: {not_in_excel[:10]}...")

import_data = []

for herb in herb_data_list:
    herb_name = herb.get("药材名称", "")
    
    if herb_name not in existing_images:
        continue
    
    image_filename = f"{herb_name}.jpg"
    image_path = os.path.join(herbs_folder, image_filename)
    
    if not os.path.exists(image_path):
        image_filename = f"{herb_name}.png"
        image_path = os.path.join(herbs_folder, image_filename)
    
    category = herb.get("分类", "")
    views = herb.get("浏览量", 0)
    if views is None:
        views = 0
    
    import_item = {
        "name": herb_name,
        "latin_name": herb.get("拉丁学名", ""),
        "medicinal_part": herb.get("药用部位", ""),
        "taste": herb.get("性味", ""),
        "meridian": herb.get("归经", ""),
        "efficacy": herb.get("功效", ""),
        "indication": herb.get("主治", ""),
        "category": category,
        "views": views,
        "images": [
            {
                "name": image_filename,
                "extname": "jpg",
                "url": f"/herbs/{image_filename}",
                "path": image_path,
                "size": os.path.getsize(image_path) if os.path.exists(image_path) else 0,
                "mimeType": "image/jpeg"
            }
        ],
        "create_date": None
    }
    
    import_data.append(import_item)

print(f"\n准备导入的药材数量: {len(import_data)}")

with open(output_path, 'w', encoding='utf-8') as f:
    json.dump(import_data, f, ensure_ascii=False, indent=2)

print(f"\n数据已保存到: {output_path}")

print("\n" + "="*50)
print("数据导入说明:")
print("="*50)
print("1. 将 herbs 文件夹中的所有图片上传到 uniCloud 云存储的 /herbs 目录")
print("2. 在 HBuilderX 中打开 uniCloud 控制台")
print("3. 进入数据库 -> yaocai 集合")
print("4. 点击'导入数据'，选择刚才生成的 JSON 文件")
print("5. 导入时注意图片字段格式，需要符合 uni-file-picker 的格式要求")
print("="*50)
